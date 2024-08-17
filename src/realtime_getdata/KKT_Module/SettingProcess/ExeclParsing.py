import os.path
import sys
import time
import openpyxl as pxl
import re
import json
import copy
from KKT_Module.ksoc_global import kgl

class KsocExcelParser():
    _DetectSymbolChar = '$'
    _RevisionSymbol = 'Revision'
    _VersionSymbol = 'Version'
    _KeySymbol = 'Key'
    _BaseAddrSymbol = 'base'
    _Reg32Symbol = 'reg32'     #registor with num bits of reg size.
    _IgnoreRegSymbol = 'ignore_reg'
    _IgnoreFieldSymbol = 'ignore'
    _UnsignedIntSymbol = 'UNSG'     #with num bits of size.
    _SignedIntSymbol = 'SIGN'     #with num bits of size.
    _FilePathSymbol = 'fpath'     #filepath
    _FileNameSymbol = 'fname'     #filename

    _offsetSymbol = 'Address_Offset'
    _bitSymbol = 'Bit'
    _bitSizeSymbol = 'BitSize'
    _nameSymbol = 'Name'
    _valueSymbol = 'Value'
    _TotalNumOfTitleSymbols = 5

    # FunctionHeadStr = 'Register_Write'
    _BaseAddrHeadStr = 'Base_Address'

    ParamSheet = {}
    file_path = ''

    @classmethod
    def parsing(cls, file_path):
        cls.file_path = file_path
        print('Excel file path : {}'.format(cls.file_path))
        cls.ParamSheet= {}
        cls.__ksocXlsxParsing()
        return cls.ParamSheet

    @classmethod
    def __ksocXlsxParsing(cls):
        cls.ParamSheet["file_name"] = cls.file_path
        wb = pxl.load_workbook(cls.file_path)
        # print(wb.sheetnames)
        # sheetnames = wb.get_sheet_names()
        sheetnames = wb.sheetnames
        sheetnameary = []
        cls.ParamSheet["workbook"] = {}
        for sheet in sheetnames:
            sname = sheet.strip()
            pos = sname.find('$')
            if pos == 0:
                sheetnameary.append(sname)
                sheet = wb[sname]
                # print(sheet.title)
                cls.ParamSheet["workbook"][sname] = {}
                cls.__sheetParsing(sheet)

        cls.ParamSheet["sheet_names"] = sheetnameary
        wb.close()

    @classmethod
    def __sheetParsing(cls, sheet):

        keyColumn = 0
        offsetColumn = 0
        bitColumn = 0
        bitSizeColumn = 0
        nameColumn = 0
        valueColumn = 0
        isKeySymbolFound = False
        param_sheet = {}
        reg_list = []
        reg_list_d = {}

        if sheet.title.strip().find("$RevisionHistory") == 0:
            ver = 0
            for column in sheet.columns:
                if column[0].value.find(cls._DetectSymbolChar + cls._VersionSymbol) >= 0:
                    for cell in column:
                        v = strToFloat(cell.value)
                        if v != None and v > ver:
                            ver = v
                    break
            param_sheet[cls._VersionSymbol] = ver
            cls.ParamSheet["workbook"][sheet.title] = param_sheet
        elif sheet.title.strip().find("$AI_WeightData") == 0:
            wfilename = []
            for row in sheet.rows:
                if row[0].value != None:
                    if row[0].value.find(cls._DetectSymbolChar + cls._FilePathSymbol) >= 0:
                        ss = row[1].value.strip()
                        param_sheet["FilePath"] = ss
                    elif row[keyColumn].value.find(cls._DetectSymbolChar + cls._FileNameSymbol) >= 0:
                        ss = row[1].value.strip()
                        if ss == None:
                            raise Exception("[ERROR] No file name, sheetname = {0}".format(sheet.title))
                        wfilename.append(ss)
            param_sheet["FileName"] = wfilename
            cls.ParamSheet["workbook"][sheet.title] = param_sheet
        else:
            idx = 0
            # find col idx
            for column in sheet.columns:
                if column[0].value == None:
                    continue
                if column[0].value.find(cls._DetectSymbolChar + cls._KeySymbol) >= 0:
                    keyColumn = idx
                    isKeySymbolFound = True
                elif column[0].value.find(cls._DetectSymbolChar + cls._offsetSymbol) >= 0:
                    offsetColumn = idx
                elif column[0].value.find(cls._DetectSymbolChar + cls._bitSizeSymbol) >= 0:
                    bitSizeColumn = idx
                elif column[0].value.find(cls._DetectSymbolChar + cls._bitSymbol) >= 0:
                    bitColumn = idx
                elif column[0].value.find(cls._DetectSymbolChar + cls._nameSymbol) >= 0:
                    nameColumn = idx
                elif column[0].value.find(cls._DetectSymbolChar + cls._valueSymbol) >= 0:
                    valueColumn = idx
                idx += 1

            if not isKeySymbolFound:
                raise Exception("[ERROR] not find Key Symbol, sheetname = {0}".format(sheet.title))

            reg_content_list = []
            enable_store_reg = False
            reg_addr_oft = None
            reg_name = None
            for row in sheet.rows:
                if row[keyColumn].value != None:
                    if row[keyColumn].value.find(cls._DetectSymbolChar + cls._BaseAddrSymbol) >= 0:
                        ba = getBaseAddr(row[offsetColumn].value.strip())
                        if ba != None:
                            param_sheet["BaseAddressStr"] = "0x{0:08X}".format(ba)
                            param_sheet["BaseAddress"] = ba
                    elif row[keyColumn].value.find(cls._DetectSymbolChar + cls._FilePathSymbol) >= 0:
                        ss = row[offsetColumn].value.strip()
                        param_sheet["FilePath"] = ss
                    elif row[keyColumn].value.find(cls._DetectSymbolChar + cls._FileNameSymbol) >= 0:
                        ss = row[offsetColumn].value.strip()
                        param_sheet["FileName"] = ss
                    elif row[keyColumn].value.find(cls._DetectSymbolChar + cls._Reg32Symbol) >= 0:
                        if enable_store_reg:
                            # print(reg_name, reg_addr_oft, param_sheet["BaseAddress"], reg_content_list)
                            rlist = cls.__convertRegContent(reg_name, reg_addr_oft, param_sheet["BaseAddress"], reg_content_list)
                            reg_list.append(rlist)
                            reg_list_d[reg_name] = reg_content_list
                        reg_addr_oft = int(row[offsetColumn].value.strip(), base=16)
                        reg_name = row[nameColumn].value.strip()
                        enable_store_reg = True
                        reg_content_list = []
                    elif row[keyColumn].value.find(cls._DetectSymbolChar + cls._IgnoreRegSymbol) >= 0:
                        if enable_store_reg:
                            rlist = cls.__convertRegContent(reg_name, reg_addr_oft, param_sheet["BaseAddress"], reg_content_list)
                            reg_list.append(rlist)
                            reg_list_d[reg_name] = reg_content_list
                        enable_store_reg = False
                        reg_addr_oft = None
                        reg_name = None
                        reg_content_list = []
                    elif row[keyColumn].value.find(cls._DetectSymbolChar + cls._UnsignedIntSymbol) >= 0:
                        if enable_store_reg:
                            if row[valueColumn].value == None:
                                print("Row's value empty sheet={}, reg_name={}, bit={}, name={}".format(sheet.title.strip(),
                                                                                                        reg_name,
                                                                                                        row[bitColumn].value,
                                                                                                        row[nameColumn].value))
                                enable_store_reg = False
                                continue
                            reg_content_list.append([row[keyColumn].value, row[bitColumn].value, row[bitSizeColumn].value, row[valueColumn].value, row[nameColumn].value])
                    elif row[keyColumn].value.find(cls._DetectSymbolChar + cls._SignedIntSymbol) >= 0:
                        if enable_store_reg:
                            if row[valueColumn].value == None:
                                print("Row's value empty sheet={}, reg_name={}, bit={}, name={}".format(sheet.title.strip(),
                                                                                                        reg_name,
                                                                                                        row[bitColumn].value,
                                                                                                        row[nameColumn].value))
                                enable_store_reg = False
                                continue
                            reg_content_list.append([row[keyColumn].value, row[bitColumn].value, row[bitSizeColumn].value, row[valueColumn].value, row[nameColumn].value])

            if enable_store_reg:
                rlist = cls.__convertRegContent(reg_name, reg_addr_oft, param_sheet["BaseAddress"], reg_content_list)
                reg_list.append(rlist)
                reg_list_d[reg_name] = reg_content_list

        param_sheet["Registers"] = reg_list
        param_sheet["Registers_d"] = reg_list_d
        cls.ParamSheet["workbook"][sheet.title] = param_sheet

    @classmethod
    def __convertRegContent(cls, reg_name, reg_adr_oft, base_addr, content_list):
        # print(reg_name, reg_adr_oft)
        # check bit field
        regex = re.compile(r'\d+')
        cvt_by_bit_pos = False
        cnt_list = []
        reg_val = 0
        for cnt in content_list:
            bit_pos_size = None
            bit_size = None
            bit_match = None
            val = None
            if cnt[1] == None and cnt[2] == None:
                raise Exception("[ERROR] convert bit fail: both bit size None, reg_name={}, addr offset={}, bit field={}".format(reg_name, reg_adr_oft, cnt[4]))
            if cnt[1] != None:
                # match = regex.match(cnt[1])
                bit_match = regex.findall(cnt[1])
                # bit_pos_size = (int(bit_match[0]) - int(bit_match[1]))
                bit_pos_size = [int(bit_match[0]), int(bit_match[1]),  (int(bit_match[0]) - int(bit_match[1]))]
                if len(bit_match) != 2 or bit_pos_size[2] < 0:
                    raise Exception("[ERROR] convert bit fail: bit size error, reg_name={}, addr offset={}, bit field={}".format(reg_name, reg_adr_oft, cnt[4]))
                cvt_by_bit_pos = True
            if cnt[2] != None:
                if not isinstance(cnt[2], int):
                    raise Exception("[ERROR] convert bit fail: bit_size is not intrger, reg_name={}, addr offset={}, bit field={}".format(reg_name, reg_adr_oft, cnt[4]))
                bit_size = cnt[2]
                # if isinstance(cnt[2], str):
                #     bitsize2 = int(cnt[2])
                # elif isinstance(cnt[2], int):
                #     bitsize2 = cnt[2]
            if bit_pos_size != None and bit_size != None:
                if (bit_pos_size[2] + 1) != bit_size:
                    raise Exception("[ERROR] convert bit fail: bit size error, reg_name={}, addr offset={}, bit field={}".format(reg_name, reg_adr_oft, cnt[4]))

            if bit_size == None:
                bit_size = bit_pos_size[2] + 1
            elif bit_size == 0:
                raise Exception("[ERROR] convert bit fail, bit size = 0, reg_name={}, addr offset={}, bit field={}".format(reg_name, reg_adr_oft, cnt[4]))

            if cnt[3] == None:
                raise Exception("[ERROR] convert value fail, reg_name={}, addr offset={}, bit field={}".format(reg_name, reg_adr_oft, cnt[4]))
            elif not isinstance(cnt[3], int):
                raise Exception("[ERROR] convert bit fail: bit size not intrger, reg_name={}, addr offset={}, bit field={}".format(reg_name, reg_adr_oft, cnt[4]))
            val = cnt[3]

            if cnt[0].find(cls._DetectSymbolChar + cls._UnsignedIntSymbol) >= 0:
                if val < 0:
                    raise Exception("[ERROR] convert value fail: unsigned value out of range, reg_name={}, addr offset={}, bit field={}".format(reg_name, reg_adr_oft, cnt[4]))
                if val > (2**bit_size) - 1:
                    raise Exception("[ERROR] convert value fail: unsigned value out of range, reg_name={}, addr offset={}, bit field={}".format(reg_name, reg_adr_oft, cnt[4]))
            elif cnt[0].find(cls._DetectSymbolChar + cls._SignedIntSymbol) >= 0:
                if bit_size == 1 and val >= 1:
                    raise Exception("[ERROR] convert value fail: signed value out of range, reg_name={}, addr offset={}, bit field={}".format(reg_name, reg_adr_oft, cnt[4]))
                if bit_size > 1 and val > (2**(bit_size - 1)) - 1 or abs(val) > (2**(bit_size - 1)):
                    raise Exception("[ERROR] convert value fail: signed value out of range, reg_name={}, addr offset={}, bit field={}".format(reg_name, reg_adr_oft, cnt[4]))

           # FOR RDI VALUE FLAG
           #  if (cnt[4] == 'RAW_MODE_B') and (cnt[1] == '[5:5]'):
           #      if kgl.raw_flag == 'RawData':
           #          val = 0
           #      elif kgl.raw_flag == 'RDIData':
           #          val = 1
           #      print("val=",val)

            cnt_list.append([cnt[0], bit_pos_size, bit_size, val])
        if cvt_by_bit_pos:
            reg_val = cls.__convertRegByBitpos(reg_name, reg_adr_oft, cnt_list)
        else:
            reg_val = cls.__convertRegByBbitsize(reg_name, reg_adr_oft, cnt_list)
        # print("addr offset=0x{:08X}, val = 0x{:X}            reg_name={}".format(reg_adr_oft + baseaddr, reg_val, reg_name))
        return [reg_name, reg_adr_oft + base_addr, reg_val, reg_adr_oft, "0x{0:08X}".format(reg_adr_oft + base_addr), "0x{0:08X}".format(reg_val)]

    @classmethod
    def __convertRegByBbitsize(cls, reg_name, reg_adr_oft, content_list):
        reg_val = 0
        bit_shift = 0
        val = 0
        for cnt in content_list:
            val = 0
            if cnt[0].find(cls._DetectSymbolChar + cls._UnsignedIntSymbol) >= 0:
                mask = 2**cnt[2] - 1
                val = cnt[3] & mask
            elif cnt[0].find(cls._DetectSymbolChar + cls._SignedIntSymbol) >= 0:
                mask = 2**cnt[2] - 1
                val = cnt[3] & mask

            val <<= bit_shift
            reg_val |= val
            bit_shift += cnt[2]
        return reg_val & 0xffffffff

    @classmethod
    def __convertRegByBitpos(cls, reg_name, reg_adr_oft, content_list):
        # print (content_list)
        # [['$UNSG32', [7, 0, 7], 8, 0]
        reg_val = 0
        for cnt in content_list:
            mask = 2**(cnt[1][2] + 1) - 1
            val = cnt[3] & mask
            val <<= cnt[1][1]
            reg_val |= val
            # if cnt[0].find(_DetectSymbolChar + _UnsignedIntSymbol) >= 0:
            #     val = cnt[2] & mask
            #     val <<= cnt[1][1]
            #     reg_val |= val
            # elif cnt[0].find(_DetectSymbolChar + _SignedIntSymbol) >= 0:
            #     val = cnt[2] & mask

        return reg_val & 0xffffffff

class ParamDictGenerator():
    class K60168:
        Sheets_map={"RFIC S2P Set Process"       :'$RFIC_S2P',
                    "Adc_Mux Parameters"         :'$Adc_MUX',
                    "Tracking Parameters"        :'$Tracking',
                    "AIACC_MEM Parameters"       :'$AIACC_MEM',
                    "AIACC_Layer Parameters"     :'$AIACC_Layer',
                    "AIACC_PARAM Parameters"     :'$AIACC_PARAM',
                    "AIACC_Siamese Parameters"   :'$AIACC_Siamese',
                    "DSPRx20M_Unit_0 Parameters" :'$DSPRx20M_Unit_0',
                    "DSPRx625K_Unit_0 Parameters":'$DSPRx625K_Unit_0',
                    "DSPRx20M_Unit_1 Parameters" :'$DSPRx20M_Unit_1',
                    "DSPRx625K_Unit_1 Parameters":'$DSPRx625K_Unit_1',
                    "DSP_Motion Parameters"      :'$DSP_Motion',
                    }
    def __init__(self, std_param_dict_path=(kgl.KKTConfig + r'/HW_setting.json')):
        '''
        Generate hardware setting parameter dictionary
        '''
        assert os.path.isfile(std_param_dict_path)
        self._std_ParamDict_path = std_param_dict_path
        self._std_ParamDict = {}
        self._current_ParamDict = {}
        self._current_ParamDict_with_val = {}
        self._std_version = ''
        self._current_version = ''
        self.initGenerator(self._std_ParamDict_path)
        pass

    def initGenerator(self, std_ParamDict_path=None):
        if std_ParamDict_path is not None:
            assert os.path.isfile(std_ParamDict_path)
            self._std_ParamDict_path = std_ParamDict_path
        print(self._std_ParamDict_path)
        self._std_ParamDict = self.readJson(self._std_ParamDict_path)
        self._std_version = str(self._std_ParamDict['RevisionHistory']['Version'])
        self._current_ParamDict = self._std_ParamDict
        self._current_version = self._std_version

    def genParamDict(self, param_sheet):
        ''' Generate parameter dictionary from hardware excel. '''
        self._current_ParamDict = {}
        workbook = param_sheet.get('workbook')
        for k, v in workbook.items():
            if k == '$RevisionHistory':
                version = v.get('Version')
                if version is None:
                    print('version empty')
                self._current_ParamDict[k.split('$')[1]] = {'Version': version}
            elif k == '$AI_WeightData':
                continue
            else:
                self._current_ParamDict[k.split('$')[1]] = {}
                registers = workbook[k]['Registers']
                for register in registers:
                    names = workbook[k]['Registers_d'][register[0]]
                    Name = []
                    bitmap = []
                    Sign = []
                    val = {}
                    pos = 0
                    for name in names:
                        Name.append(name[4].strip())
                        Sign.append(name[0].strip())
                        if (type(pos) and type(name[2])) == int:
                            bitmap.append([name[2], pos])
                            pos = pos + name[2]
                        elif type(name[1]) == str:
                            # print(register)
                            # print(name)
                            pos = eval(name[1].replace(':', ','))[1]
                            size = eval(name[1].replace(':', ','))[0] - pos + 1
                            bitmap.append([size, pos])
                        # else:
                        #     print(register[0], name[4], pos, name[2])
                    self._current_ParamDict[k.split('$')[1]][str(register[4])] = [register[0], bitmap, Name, Sign, val]
        self._current_version = str(self._current_ParamDict['RevisionHistory']['Version'])
        return self._current_ParamDict

    def writeJson(self, dict=None, dictname=None):
        '''
           Write parameter dictionary to json file parse from hardware excel.
        '''
        if dictname is None:
            dictname = self._std_ParamDict_path
        if dict is None:
            assert self._current_ParamDict != {}, 'None cuurent param dict'
            dict = self._current_ParamDict
        with open(dictname , 'w') as outfile:
            json.dump(dict, outfile, ensure_ascii=False)
            outfile.write('\n')
        print('dict save')

    def readJson(self, path):
        ''' Write parameter dictionary to json file parse from hardware excel.'''
        with open(path, newline='') as json_file:
            param_dict = json.load(json_file)
        print('get dict')
        return param_dict

    def writeRegVal(self, read_procList, ParamDict=None):
        '''
           Write register value from process list to parameter dictionary.
        '''
        # sheets = list(ParamDict.keys())
        if ParamDict is None:
            ParamDict = copy.deepcopy(self._std_ParamDict)
        write = False
        sheet_name =''
        for l in read_procList:
            # print(l)
            if l[0] == 'Remark':
                sheet_name = self.K60168.Sheets_map.get(l[1])
                if sheet_name is None:
                    write = False
                    continue
                sheet_name = sheet_name.split('$')[1]
                write = True
            elif l[0] == "Register" and write:
                # if l[1]==1074462852:
                #     s=2
                addr = hex(l[1]).zfill(10).upper().replace('X', 'x')
                bitmap = ParamDict[sheet_name][addr][1]
                if type(ParamDict[sheet_name][addr][3]) is dict:
                    signlist = ParamDict[sheet_name][addr][4]
                else:
                    signlist = ParamDict[sheet_name][addr][3]
                val_list = val2vallist(bitmap, l[2], signlist)
                # print(addr, bitmap, signlist, val_list)
                if len(val_list) == len(ParamDict[sheet_name][addr][2]):
                    nameval = dict(zip(ParamDict[sheet_name][addr][2], val_list))
                    ParamDict[sheet_name][addr].insert(3, nameval)
                else:
                    print(ParamDict[sheet_name][addr], val_list, 'field not match')
            else:
                write = False
        self._current_ParamDict_with_val = ParamDict
        return ParamDict



def readProcListFromFile(filename):
    read_procList =[]
    regex = re.compile(r'0x[0-9A-Fa-f]+')
    with open(filename) as in_file:
        for line in in_file:
            if line.find(getSymbolString('RegSymbol')) >= 0:
                val = regex.findall(line)
                if len(val) >= 2:
                    read_procList.append(['RegSymbol', int(val[0], 16), int(val[1], 16)])
            elif line.find(getSymbolString('CommentSymbol')) >= 0:
                if line.find(getSymbolString('RFFileSymbol')) >= 0:
                    val = line.split(':')
                    read_procList.append(['RFFileSymbol', val[1].strip()])
                elif line.find(getSymbolString('AIWeightPathSymbol')) >= 0:
                    val = line.split(':')
                    read_procList.append(['AIWeightPathSymbol', val[1].strip()])
                elif line.find(getSymbolString('AIWeightFilesSymbol')) >= 0:
                    val = line.split(':')
                    files = val[1].strip().split(' ')
                    read_procList.append(['AIWeightFilesSymbol', files])
                else:
                    val = line.replace(getSymbolString('CommentSymbol'), "")
                    read_procList.append(['CommentSymbol', val.strip()])

    return read_procList

def getSymbolString(symbol):
    return{
        'RegSymbol'           : 'reg_write',
        'CommentSymbol'       : '//',
        # kgl.WeightFileSymbol    : '// WeightFile',
        'RFFileSymbol'        : '// RFFile',
        'AIWeightPathSymbol'  : '// AIWeightPath',
        'AIWeightFilesSymbol' : '// AIWeightFiles',
    }.get(symbol,None)

def val2vallist(bitsMap, val, signlist):
    val_list = []
    for i in range(len(bitsMap)):
        if signlist[i] in ['$UNSG32', '$UNSG33'] :
            mask = 2 ** bitsMap[i][0] - 1
            new_val =  val >> bitsMap[i][1]
            v = new_val & mask
            val_list.append(v)
        elif signlist[i] == '$SIGN32':
            mask = 2 ** (bitsMap[i][0]) - 1
            new_val =  val >> bitsMap[i][1]
            v = new_val & mask
            sign = bin(v)[2:].zfill(bitsMap[i][0])[0]
            v = int('0b'+bin(v)[2:].zfill(bitsMap[i][0])[1:], 2)
            if sign == '0':
                val_list.append(v)
            elif sign == '1':
                v = -((v ^ (2 ** (bitsMap[i][0]-1)) - 1) +1)
                val_list.append(v)
    return  val_list

def getBaseAddr(addr):
    return{
        'TRK_BA'            :0x50000500,
        'DSPRx20M_Unit_0'   :0x400D0000,
        'DSPRx625K_Unit_0'  :0x400B0000,
        'DSPRx20M_Unit_1'   :0x400F0000,
        'DSPRx625K_Unit_1'  :0x40090000,
        'DSP_Motion'        :0x4005C000,
        'AIACC'             :0x40060000,
        'GCR_BA'            :0x50000000,
        'CLK_BA'            :0x50000200,
        'SPI_RFIC_BA'       :0x400A0000,
        'AI_WEIGHT_BA'      :0x20020000,
    }.get(addr,None)

def strToInt(s):
    try:
        return int(s)
    except ValueError:
        return None

def strToFloat(s):
    try:
        f = float(s)
        return f
    except :
        return None

if __name__ == '__main__':
    from KKT_Module.ksoc_global import kgl
    s = time.time()
    file_name = r"C:\Users\eric.li\Desktop\Python\0_Standard_Ksoc_Tool\TempParam\K60168-Develop-60000-v1.2.3-20210512-bin_new\K60SoCB_MPL_v1_10.xlsx"
    p = KsocExcelParser.parsing(file_name)
    KsocHWSetting = ParamDictGenerator(os.path.join(kgl.KKTConfig, 'HW_setting.json'))
    KsocHWSetting.genParamDict(p)
    KsocHWSetting.writeJson()

    pd = KsocHWSetting.readJson(os.path.join(kgl.KKTConfig, 'HW_setting.json'))
    proc_list = readProcListFromFile(
        r'C:\Users\eric.li\Desktop\Python\KKT_Python_Module\TempParam\210906_SIAMESE_ON\param\param_210628_KSOCB_SIAMESE_ON.txt')

    pd_w = KsocHWSetting.writeRegVal(read_procList=proc_list)

    print(time.time() - s)
    input('any:')









